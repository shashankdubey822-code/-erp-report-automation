from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .utilities import format_attendance_value

# -------------------------
# Excel generation (openpyxl)
# -------------------------
def get_adjusted_width(worksheet, column_cells) -> int:
    """Calculate a reasonable width for an openpyxl column based on text length."""
    max_length = 0
    for c in column_cells:
        try:
            if c.value is not None:
                length = len(str(c.value))
                if length > max_length:
                    max_length = length
        except Exception:
            pass
    # small fudge factor
    return max(8, min(50, max_length + 2))


def update_worksheet(worksheet, df: pd.DataFrame):
    """
    Auto-adjust column widths on an openpyxl worksheet.
    For subject columns we keep a fixed width to avoid very wide sheets.
    """
    for i, col in enumerate(worksheet.columns, start=1):
        col_letter = get_column_letter(i)
        # For subject columns we set a reasonable width, else auto adjust
        # We treat subject columns as those after 3rd and before last 3 columns
        try:
            if 3 < i <= df.shape[1] - 3:
                worksheet.column_dimensions[col_letter].width = 15
            else:
                worksheet.column_dimensions[col_letter].width = get_adjusted_width(worksheet, col)
        except Exception:
            # fallback
            worksheet.column_dimensions[col_letter].width = 15


def add_custom_header(ws, metadata):
    """Adds a dynamic, multi-line, centered, and bolded header to the worksheet."""
    # Safe fallbacks for metadata
    dept_name = metadata.get('department_name', 'DEPT OF COMPUTER SCIENCE & TECHNOLOGY')
    academic_year = metadata.get('academic_year', '2025-2026')
    semester = metadata.get('semester', 'Odd')
    report_title = metadata.get('report_title', 'ATTENDANCE MONITORING REPORT')
    branch = metadata.get('branch', 'MRU-School of Engineering')
    department_specialization = metadata.get('department_specialization', 'B.Tech (Hons.) in Computer Science Engineering with specializations in Gen AI')
    class_name_division = metadata.get('class_name_division', 'B.Tech CSE Gen AI Sem 1 | Division: All')
    date_range = metadata.get('date_range', '28/07/2025 to 19/09/2025 (2025-2026)')
    coordinator = metadata.get('coordinator', '')

    # Define header lines
    header_lines = [
        (dept_name, 'A1'),
        (f"Academic Year: {academic_year} - Semester: {semester}", 'A2'),
        (report_title, 'A3'),
        (f"Branch: {branch}", 'A4'),
        (f"Department: {department_specialization}", 'A5'),
        (f"Class Name: {class_name_division}", 'A6'),
        (f"Date: {date_range}", 'A7'),
        (f"Program Coordinator: {coordinator}", 'A8')
    ]

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    for line, cell_ref in header_lines:
        cell = ws[cell_ref]
        cell.value = line
        cell.font = bold_font
        cell.alignment = center_alignment
        # Merge cells from column A to the last column of the header
        ws.merge_cells(start_row=cell.row, start_column=1, end_row=cell.row, end_column=8)

def create_excel_file(df, subject_details, metadata, chart_image=None, report_color='#FFFF00', subjects_with_zero_attendance: list = None):
    """Create an excel file safely without corrupting fills."""

    if "Roll No_duplicate" in df.columns and "Roll No" not in df.columns:
        df = df.rename(columns={"Roll No_duplicate": "Roll No"})

    wb = Workbook()
    ws = wb.active
    ws.title = metadata.get("monitoring_stage", "Report")
    
    # Add the custom header
    add_custom_header(ws, metadata)

    # Determine the starting row for the main data based on the custom header
    # and potentially a warning message for zero attendance subjects.
    start_row = 10 

    if subjects_with_zero_attendance:
        zero_att_message = "The following subjects have 0% attendance for all students: " + \
                           ", ".join(subjects_with_zero_attendance) + \
                           ". These subjects are not included in the main table."
        
        # Add the warning message
        ws.cell(row=start_row, column=1, value=zero_att_message).font = Font(bold=True, color="FF0000") # Red color for warning
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(df.columns) + 2) # Merge across all expected columns
        start_row += 2 # Add some space after the warning
    
    # Styles (valid only)
    vibrant_fill = PatternFill(start_color=report_color.replace("#", ""), end_color=report_color.replace("#", ""), fill_type="solid")
    grey_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    light_blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    # Header rows
    headers = list(df.columns)

    overall_idx = (headers.index("Overall %age of all subjects from ERP report") + 1
                   if "Overall %age of all subjects from ERP report" in headers else 3)

    # Write header row
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if col_index <= overall_idx:
            cell.fill = vibrant_fill

    # Write data
    min_attendance = metadata.get("min_attendance", 75)
    for r, row_data in enumerate(df.values.tolist(), start=start_row + 1):
        for c, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

            # Highlight subject columns
            if 4 <= c <= df.shape[1] - 3:
                formatted_val = format_attendance_value(val)
                cell.value = formatted_val
                try:
                    if float(val) < min_attendance:
                        cell.fill = grey_fill
                except:
                    pass

    # Summary table
    summary_start = start_row + len(df) + 3
    subjects = list(subject_details.keys())
    valid_subjects = [s for s in subjects if s in df.columns]

    summary_headers = ["Subject", "<75%", "<70%", "<65%", "<60%"]

    for i, header in enumerate(summary_headers, start=2):
        cell = ws.cell(row=summary_start, column=i, value=header)
        cell.font = Font(bold=True)
        cell.fill = light_blue_fill
        cell.border = thin_border

    for r, subject in enumerate(valid_subjects, start=summary_start + 1):
        ws.cell(row=r, column=2, value=subject).fill = light_blue_fill
        thresholds = [75, 70, 65, 60]
        for j, t in enumerate(thresholds, start=3):
            count = (df[subject] < t).sum()
            cell = ws.cell(row=r, column=j, value=int(count))
            cell.fill = light_blue_fill
            cell.border = thin_border

    # Autosize
    update_worksheet(ws, df)

    # Chart image
    if chart_image:
        img = XLImage(chart_image)
        img.anchor = f"B{summary_start + len(valid_subjects) + 3}"
        ws.add_image(img)

    # Save workbook
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
