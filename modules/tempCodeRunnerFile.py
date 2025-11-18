import csv
import logging
from io import BytesIO, StringIO

import pandas as pd
from matplotlib.figure import Figure
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def create_report_dataframe(erp_file, min_attendance_criteria):
    """Create a dataframe from the ERP file."""
    erp_file.seek(0)
    file_content = erp_file.read().decode('utf-8', errors='ignore')

    string_io_file = StringIO(file_content)

    reader = csv.reader(string_io_file)
    lines_as_lists = list(reader)
    lines = [",".join(row) for row in lines_as_lists]

    header_start_index = -1

    header_patterns = [
        'Sr.,Division/Section,Unique id',
        'Sr,Division/Section,Unique id',
        'Sr.,Division,Unique id',
        'Unique id',
        'Roll',
        'Student Name'
    ]

    for i, line in enumerate(lines):
        line_clean = line.replace('"', '').replace(' ', '')
        for pattern in header_patterns:
            pattern_clean = pattern.replace(' ', '')
            if pattern_clean in line_clean:
                header_start_index = i
                logger.info("Found header at line %d with pattern: %s", i, pattern)
                break
        if header_start_index != -1:
            break

    if header_start_index == -1:
        logger.error("Could not find data table header. Searched for patterns: %s", header_patterns)
        logger.error("First 10 lines of file: %s", lines[:10])
        raise ValueError("Could not find the data table header in the ERP file. Please check the file format.")

    try:
        h1_subjects = [h.strip() for h in lines_as_lists[header_start_index]]

        max_lines = len(lines_as_lists)
        h2_codes = [h.strip() for h in lines_as_lists[header_start_index + 2]] if header_start_index + 2 < max_lines else ['']
        h3_types = [h.strip() for h in lines_as_lists[header_start_index + 3]] if header_start_index + 3 < max_lines else ['']
        h_metrics = [h.strip() for h in lines_as_lists[header_start_index + 4]] if header_start_index + 4 < max_lines else ['']

    except IndexError as e:
        logger.error("Error parsing headers at index %d: %s", header_start_index, e)
        raise ValueError(f"Invalid file format: Unable to parse headers starting at line {header_start_index + 1}") from e

    last_subject, last_code, last_type = "", "", ""
    for i, subject in enumerate(h1_subjects):
        if subject:
            last_subject = subject
        else:
            h1_subjects[i] = last_subject
        if h2_codes[i]:
            last_code = h2_codes[i]
        else:
            h2_codes[i] = last_code
        if h3_types[i]:
            last_type = h3_types[i]
        else:
            h3_types[i] = last_type

    final_headers, subject_details = [], {}
    for i, metric in enumerate(h_metrics):
        subject, code, subject_type = h1_subjects[i], h2_codes[i], h3_types[i]
        if subject in ['Sr.', 'Division/Section', 'Unique id', 'Rollno', 'Student Name', 'PRN / Enroll']:
            final_headers.append(subject)
        elif 'Total' in subject:
            final_headers.append(f"Grand Total - {metric}")
        else:
            final_headers.append(f"{subject} - {metric}")
            if subject not in subject_details:
                subject_details[subject] = {'code': code, 'type': subject_type}

    data_as_string_list = [",".join(row) for row in lines_as_lists[header_start_index + 6:]]
    csv_data_string = "\n".join(data_as_string_list)

    df = pd.read_csv(StringIO(csv_data_string), header=None, names=final_headers, on_bad_lines='skip')
    df.dropna(subset=['Rollno'], inplace=True)

    output_df = pd.DataFrame({'Sr No.': range(1, len(df) + 1), 'Roll No': df['Rollno'], 'Student Name': df['Student Name']})

    subject_percent_cols = {}
    for subject in subject_details:
        found_column = None
        possible_suffixes = [' - Total %', ' - % (PP)', ' - % (PR)', ' - % (TUT)', ' - %', ' - Total', '- Total %', '- % (PP)',
                             '- % (PR)', '- % (TUT)']

        for suffix in possible_suffixes:
            potential_col = f"{subject}{suffix}"
            if potential_col in df.columns:
                found_column = potential_col
                break

        if found_column:
            subject_percent_cols[subject] = found_column
        else:
            logger.warning("No matching column found for subject: %s", subject)

    for subject, col_name in subject_percent_cols.items():
        output_df[subject] = pd.to_numeric(df[col_name], errors='coerce').fillna(0)

    output_df['Overall %age of all subjects from ERP report'] = pd.to_numeric(
        df.get('Grand Total - %', df.get('Total - %', 0)), errors='coerce').fillna(0)
    output_df['Roll No_duplicate'] = df['Rollno']

    subject_keys = list(subject_percent_cols.keys())

    if subject_keys:
        output_df['Count of Courses with attendance below minimum attendance criteria'] = output_df[subject_keys].apply(
            lambda row: (row < min_attendance_criteria).sum(), axis=1)
    else:
        output_df['Count of Courses with attendance below minimum attendance criteria'] = 0

    output_df['Whether Critical'] = output_df['Count of Courses with attendance below minimum attendance criteria'].apply(
        lambda count: 'CRITICAL' if count > 4 else 'Not Critical')

    return output_df, subject_details


def create_excel_file(df, subject_details, metadata, chart_image=None):
    """Create an excel file from the dataframe."""
    df.rename(columns={'Roll No_duplicate': 'Roll No'}, inplace=True)
    output_buffer = BytesIO()
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        sheet_name = metadata['monitoring_stage']

        worksheet = writer.book.create_sheet(title=sheet_name)
        if 'Sheet' in writer.book.sheetnames:
            writer.book.remove(writer.book['Sheet'])

        # Define styles
        vibrant_yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        grey_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
        light_blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        black_font = Font(color="000000", bold=True)
        summary_font = Font(color="000000", bold=True, size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Header section
        last_col_letter = get_column_letter(df.shape[1])
        title_text = (
            "DEPARTMENT OF CST\n"
            "JUL-DEC 2025\n"
            "LOW ATTENDANCE REVIEW REPORT (1 WEEK PRIOR TO LAST DAY OF CLASSES)"
        )
        cell_b2 = worksheet.cell(row=2, column=2)
        cell_b2.value = title_text
        cell_b2.font = Font(color="000000", bold=True, size=12)
        cell_b2.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        worksheet.merge_cells(f'B2:{last_col_letter}2')
        worksheet.row_dimensions[2].height = 45

        metadata_row = 4
        metadata_text = (
            f"Branch: MRU-School of Engineering\n"
            f"Department: Bachelor of Technology in Computer Science and Engineering\n"
            f"Class Name: {metadata['class_name']} | Division: {metadata['division']}\n"
            f"Date: {metadata['date_range']}\n"
            f"Program Coordinator: {metadata['coordinator']}"
        )
        cell_b4 = worksheet.cell(row=metadata_row, column=2)
        cell_b4.value = metadata_text
        cell_b4.font = Font(color="000000", bold=False, size=9)
        cell_b4.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        worksheet.merge_cells(f'B{metadata_row}:{last_col_letter}{metadata_row}')
        worksheet.row_dimensions[metadata_row].height = 60

        # Main data table
        table_start_row = 6
        headers1 = list(df.columns)
        try:
            headers2 = [''] * 3 + [subject_details.get(subj, {}).get('code', '') for subj in subject_details] + [''] * 4
            headers3 = [''] * 3 + [subject_details.get(subj, {}).get('type', '') for subj in subject_details] + [''] * 4
            overall_percent_col_index = headers1.index('Overall %age of all subjects from ERP report') + 1
        except (ValueError, KeyError):
            overall_percent_col_index = df.shape[1] - 3
            headers2 = [''] * len(headers1)
            headers3 = [''] * len(headers1)

        for i, val in enumerate(headers1, 1):
            cell = worksheet.cell(row=table_start_row, column=i, value=val)
            cell.font = black_font
            cell.border = thin_border
            if i <= overall_percent_col_index:
                cell.fill = vibrant_yellow_fill
        for i, val in enumerate(headers2, 1):
            cell = worksheet.cell(row=table_start_row + 1, column=i, value=val)
            cell.font = black_font
            cell.border = thin_border
            if i <= overall_percent_col_index:
                cell.fill = vibrant_yellow_fill
        for i, val in enumerate(headers3, 1):
            cell = worksheet.cell(row=table_start_row + 2, column=i, value=val)
            cell.font = black_font
            cell.border = thin_border
            if i <= overall_percent_col_index:
                cell.fill = vibrant_yellow_fill

        data_start_row = table_start_row + 3
        min_attendance = metadata.get('min_attendance', 75)
        for r_idx, row_data in enumerate(df.values.tolist(), data_start_row):
            for c_idx, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                cell.border = thin_border

                # Default fill
                cell.fill = vibrant_yellow_fill

                # Check for subject columns (from 4th column to the 4th last column)
                if 3 < c_idx <= df.shape[1] - 3:
                    try:
                        # Check if the value is a number and less than the minimum attendance
                        if float(value) < min_attendance:
                            cell.fill = grey_fill
                    except (ValueError, TypeError):
                        # Ignore if the value is not a number
                        pass

        # New Transposed Summary table
        summary_start_row = len(df) + data_start_row + 3
        subjects = list(subject_details.keys())
        valid_subjects = [s for s in subjects if s in df.columns]

        summary_headers = ["Subject", "Number of students below 75%", "Number of students below 70%",
                         "Number of students below 65%", "Number of students below 60%"]

        # Write summary table headers
        for i, header in enumerate(summary_headers):
            cell = worksheet.cell(row=summary_start_row, column=i + 2, value=header)
            cell.font = summary_font
            cell.fill = light_blue_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Write summary table data
        for r_idx, subject in enumerate(valid_subjects, summary_start_row + 1):
            # Subject name
            cell = worksheet.cell(row=r_idx, column=2, value=subject)
            cell.font = summary_font
            cell.fill = light_blue_fill
            cell.border = thin_border
            cell.alignment = left_align

            # Number of students below thresholds
            thresholds = [75, 70, 65, 60]
            for c_idx, threshold in enumerate(thresholds, 3):
                count = (df[subject] < threshold).sum()
                cell = worksheet.cell(row=r_idx, column=c_idx, value=int(count))
                cell.font = summary_font
                cell.fill = light_blue_fill
                cell.border = thin_border
                cell.alignment = center_align

        footer_start_row = summary_start_row + len(valid_subjects) + 2
        cell = worksheet.cell(row=footer_start_row, column=15, value="Signature of Mentor")
        cell.font = black_font

        # Auto-adjust column widths
        for col in worksheet.columns:
            if any(c.value for c in col):
                max_length = max(len(str(c.value)) for c in col if c.value)
                adjusted_width = max_length + 2

                col_letter = get_column_letter(col[0].column)
                if col_letter == 'B':  # Assuming 'Roll No' is consistently in column B
                    worksheet.column_dimensions[col_letter].width = min(adjusted_width, 15)  # Cap at 15
                else:
                    worksheet.column_dimensions[col_letter].width = adjusted_width

        # Add chart image
        if chart_image:
            img = Image(chart_image)
            img.anchor = 'B' + str(footer_start_row + 2)  # Anchor image to cell
            worksheet.add_image(img)

    output_buffer.seek(0)
    return output_buffer


def generate_summary_table_html(df, min_attendance=75):
    """Generates an HTML summary table from the dataframe."""

    subject_columns = [col for col in df.columns if col not in
                       ['Sr No.', 'Roll No', 'Student Name', 'Overall %age of all subjects from ERP report',
                        'Roll No_duplicate', 'Count of Courses with attendance below minimum attendance criteria',
                        'Whether Critical']]

    summary_data = []
    for subject in subject_columns:
        summary_data.append({
            'Subject': subject,
            f'Below {min_attendance}%': (df[subject] < min_attendance).sum(),
            'Below 70%': (df[subject] < 70).sum(),
            'Below 65%': (df[subject] < 65).sum(),
            'Below 60%': (df[subject] < 60).sum(),
        })

    summary_df = pd.DataFrame(summary_data)
    return summary_df.to_html(classes='summary-table', index=False)


def generate_chart_image(df):
    """Generates a chart image from the dataframe and returns it as a BytesIO object."""

    subject_columns = [col for col in df.columns if col not in
                       ['Sr No.', 'Roll No', 'Student Name', 'Overall %age of all subjects from ERP report',
                        'Roll No_duplicate', 'Count of Courses with attendance below minimum attendance criteria',
                        'Whether Critical']]

    courses = subject_columns
    students_below_75 = [(df[course] < 75).sum() for course in courses]

    fig = Figure(figsize=(15, 8))
    ax = fig.subplots()
    chart_bars = ax.bar(courses, students_below_75, color='skyblue')
    ax.set_title('Number of Students with Attendance Below 75% per Course', pad=20)
    ax.set_xlabel('Courses')
    ax.set_ylabel('Number of Students below 75%')
    ax.tick_params(axis='x', rotation=45, labelsize=8)
    for bar_item in chart_bars:
        yval = bar_item.get_height()
        ax.text(bar_item.get_x() + bar_item.get_width() / 2.0, yval, str(int(yval)), va='bottom')
    fig.tight_layout()

    buf = BytesIO()
    fig.savefig(buf, format="png")
    buf.seek(0)
    return buf