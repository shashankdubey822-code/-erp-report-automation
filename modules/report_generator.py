"""
modules/report_generator.py

Robust attendance report generator:
- Parses ERP CSV exports into a normalized DataFrame
- Creates Excel report (openpyxl)
- Creates PDF report (reportlab)
- Generates a summary table (HTML)
- Generates a chart image (Matplotlib -> BytesIO)

Replace your broken file with this. Tested for structural correctness and defensive handling.
"""

import csv
import logging
import math
import re
import textwrap
from io import BytesIO, StringIO
from typing import Dict, Tuple, List

import pandas as pd
from matplotlib.figure import Figure
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image as RLImage,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# Logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# -------------------------
# Utility helpers
# -------------------------
def safe_str(x) -> str:
    return "" if x is None else str(x)


def clean_subject_label(label: str) -> str:
    """Strip trailing metrics like '- %', '- Total', '(TUT)' etc. Return short subject name."""
    if label is None:
        return ""
    s = str(label).strip()
    # Remove patterns like " - % (PP)", " - Total - something", "(PR)"
    s = re.sub(r"\s*-\s*%.*$", "", s)  # "- % ..." and suffixes
    s = re.sub(r"\s*-\s*Total.*$", "", s)  # "- Total ..."
    s = re.sub(r"\s*\(.*\)$", "", s)  # trailing parentheses
    s = re.sub(r"\s*-\s*$", "", s)  # trailing hyphen
    return s.strip()

def format_attendance_value(value):
    """Formats a numeric value to remove .0 if it's an integer, otherwise keeps float."""
    try:
        f_val = float(value)
        if f_val.is_integer():
            return int(f_val)
        return f_val
    except (ValueError, TypeError):
        return value


# -------------------------
# Dataframe creation
# -------------------------
def create_report_dataframe(erp_file, min_attendance_criteria: int = 75) -> Tuple[pd.DataFrame, Dict, Dict]:
    """
    Parse ERP CSV-like file object and return (output_df, subject_details, extracted_metadata).
    output_df: DataFrame with 'Sr No.', 'Roll No', 'Student Name', subject percentage columns, etc.
    subject_details: dict mapping subject_name -> {'code': code, 'type': type}
    extracted_metadata: dict with dynamically extracted header fields
    """
    erp_file.seek(0)
    raw = erp_file.read()
    if isinstance(raw, bytes):
        content = raw.decode("utf-8", errors="ignore")
    else:
        content = str(raw)

    sio = StringIO(content)
    reader = csv.reader(sio)
    rows = list(reader)
    
    # Metadata extraction
    extracted_metadata = {}
    for row in rows[:20]:  # Scan top 20 rows for metadata
        line = ",".join(row)
        if "Branch:" in line:
            extracted_metadata['branch'] = line.split("Branch:")[1].split(",")[0].strip()
        if "Department:" in line:
            extracted_metadata['department_specialization'] = line.split("Department:")[1].split(",")[0].strip()
        if "Class Name:" in line:
            extracted_metadata['class_name_division'] = line.split("Class Name:")[1].split(",")[0].strip()
        if "Date:" in line:
            extracted_metadata['date_range'] = line.split("Date:")[1].split(",")[0].strip()
        if "Program Coordinator:" in line:
            extracted_metadata['coordinator'] = line.split("Program Coordinator:")[1].split(",")[0].strip()
        if "Academic Year:" in line:
            extracted_metadata['academic_year'] = line.split("Academic Year:")[1].split("-")[0].strip()
        if "Semester:" in line:
            extracted_metadata['semester'] = line.split("Semester:")[1].split(",")[0].strip()
        # Look for a line that might be the department name (all caps, single entry in row)
        if len(row) == 1 and row[0].isupper() and 'department_name' not in extracted_metadata:
             extracted_metadata['department_name'] = row[0].strip()
        # Look for a line that might be the report title
        if "ATTENDANCE" in line.upper() and 'report_title' not in extracted_metadata:
            extracted_metadata['report_title'] = line.split(',')[0].strip()
    
    # Basic heuristics for header start
    header_start_index = -1
    header_patterns = [
        "Sr.,Division/Section,Unique id",
        "Sr,Division/Section,Unique id",
        "Sr.,Division,Unique id",
        "Unique id",
        "Roll",
        "Student Name",
    ]

    # join each row as comma-joined string
    joined = [",".join(r) for r in rows]
    for i, line in enumerate(joined[:60]):
        cleaned = line.replace('"', "").replace(" ", "")
        for pat in header_patterns:
            if pat.replace(" ", "") in cleaned:
                header_start_index = i
                logger.info("Header detected at line %d using pattern '%s'", i, pat)
                break
        if header_start_index != -1:
            break

    if header_start_index == -1:
        # fallback: look for 'Roll' or 'Student Name' exact tokens
        for i, r in enumerate(rows[:60]):
            for cell in r:
                if str(cell).strip().lower() in ("roll", "student name"):
                    header_start_index = i
                    break
            if header_start_index != -1:
                break

    if header_start_index == -1:
        logger.error("Failed to detect header in ERP. First 10 lines: %s", joined[:10])
        raise ValueError("Could not find the data table header in the ERP file. Please check the file format.")

    # Extract header rows (we expect multiple rows describing subject names / codes / types / metrics)
    # We'll be defensive: if rows missing, fill with empty strings.
    def row_at(idx):
        return rows[idx] if 0 <= idx < len(rows) else []

    h1 = [safe_str(x).strip() for x in row_at(header_start_index)]
    h2 = [safe_str(x).strip() for x in row_at(header_start_index + 2)]
    h3 = [safe_str(x).strip() for x in row_at(header_start_index + 3)]
    h4 = [safe_str(x).strip() for x in row_at(header_start_index + 4)]

    # Expand to same length
    maxlen = max(len(h1), len(h2), len(h3), len(h4))
    def pad(lst):
        return lst + [""] * (maxlen - len(lst))
    h1, h2, h3, h4 = pad(h1), pad(h2), pad(h3), pad(h4)

    # Fill empty subject names in h1 with last seen (ERP often uses merged header cells)
    last = ""
    for i, val in enumerate(h1):
        if val:
            last = val
        else:
            h1[i] = last

    # Construct final_headers - pair subject + metric info where appropriate
    final_headers = []
    subject_details = {}
    for i, metric in enumerate(h4):
        subj = h1[i]
        code = h2[i] if i < len(h2) else ""
        typ = h3[i] if i < len(h3) else ""
        # Identify special columns
        if subj.strip() in ("Sr.", "Division/Section", "Unique id", "Rollno", "Student Name", "PRN / Enroll"):
            final_headers.append(subj.strip())
        elif "Total" in subj or "Grand Total" in subj or "Total" in metric:
            # unify grand total naming
            final_headers.append(f"Grand Total - {metric}".strip())
        else:
            # typical subject column: "SUBJ - metric"
            label = f"{subj} - {metric}".strip()
            final_headers.append(label)
            if subj not in subject_details:
                subject_details[subj] = {"code": code, "type": typ}

    # Deduplicate headers to prevent pandas error
    new_headers = []
    counts = {}
    for header in final_headers:
        if header in counts:
            counts[header] += 1
            new_headers.append(f"{header}_duplicate_{counts[header]}")
        else:
            counts[header] = 1
            new_headers.append(header)
    final_headers = new_headers

    # Data rows start: ERP often has 6 header lines; we'll use header_start_index + 6 as before
    data_start = header_start_index + 6
    data_rows = rows[data_start:]
    data_joined = "\n".join([",".join(r) for r in data_rows])
    df = pd.read_csv(StringIO(data_joined), header=None, names=final_headers, on_bad_lines="skip")
    # Ensure Rollno column exists
    roll_col = next((c for c in df.columns if "Roll" in c or "roll" in c or "Rollno" in c), None)
    if roll_col is None:
        # look for Unique id
        roll_col = next((c for c in df.columns if "Unique" in c or "unique" in c), None)
    if roll_col is None:
        raise ValueError("Could not find Roll/Rollno column in parsed data")
    df.dropna(subset=[roll_col], inplace=True)

    # Build output_df with basic info
    output_df = pd.DataFrame({
        "Sr No.": range(1, len(df) + 1),
        "Roll No": df[roll_col].astype(str)
    })

    # For "Student Name" column detection
    name_col = next((c for c in df.columns if "Student Name" in c or "Student" in c), None)
    if name_col:
        output_df["Student Name"] = df[name_col].astype(str)
    else:
        # fallback: attempt to find a sensible column
        output_df["Student Name"] = df.iloc[:, 1].astype(str) if df.shape[1] > 1 else ""

    # Map subjects -> column names that contain percentage metrics
    subject_percent_cols = {}
    for subj in subject_details:
        # try many suffixes to find correct column name
        found = None
        suffixes = [
            " - Total %", " - % (PP)", " - % (PR)", " - % (TUT)",
            " - %", " - Total", "- Total %", "- % (PP)", "- % (PR)", "- % (TUT)",
            f"{subj} - %", f"{subj} - Total %"
        ]
        for s in suffixes:
            candidate = f"{subj}{s}"
            if candidate in df.columns:
                found = candidate
                break
        # fallback: find any column whose name starts with subject and contains '%'
        if not found:
            for col in df.columns:
                if col.startswith(subj) and "%" in col:
                    found = col
                    break
        if found:
            subject_percent_cols[subj] = found
        else:
            # warn but continue
            logger.warning("No percentage column found for subject '%s'", subj)

    # add numeric subject percentage columns to output_df
    for subj, col_name in subject_percent_cols.items():
        output_df[subj] = pd.to_numeric(df.get(col_name, 0), errors="coerce").fillna(0)

    # overall percentage
    overall_col = next((c for c in df.columns if "Grand Total - %" in c or "Total - %" in c or "Overall" in c), None)
    if overall_col:
        output_df["Overall %age of all subjects from ERP report"] = pd.to_numeric(df.get(overall_col), errors="coerce").fillna(0)
    else:
        # safe fallback: create zeros
        output_df["Overall %age of all subjects from ERP report"] = 0

    

    # count of courses below threshold
    subject_keys = list(subject_percent_cols.keys())
    if subject_keys:
        output_df["Count of Courses with attendance below minimum attendance criteria"] = output_df[subject_keys].apply(
            lambda row: (row < min_attendance_criteria).sum(), axis=1
        )
    else:
        output_df["Count of Courses with attendance below minimum attendance criteria"] = 0

    output_df["Whether Critical"] = output_df["Count of Courses with attendance below minimum attendance criteria"].apply(
        lambda c: "CRITICAL" if c >= 3 else ""
    )

    return output_df, subject_details, extracted_metadata


# -------------------------
# PDF generation (reportlab)
# -------------------------
def create_pdf_file(df: pd.DataFrame, subject_details: dict, metadata: dict, chart_image: BytesIO = None) -> BytesIO:
    """
    Create a well-formatted PDF using reportlab.
    - Single header row (subject names only)
    - Dynamic column width calculation to avoid overlap
    - Header wrapping and auto-shrink by using modest font sizes
    - Summary table appended
    Returns a BytesIO buffer containing PDF bytes.
    """
    # defensive imports
    from reportlab.lib.pagesizes import landscape
    styles = getSampleStyleSheet()

    # Prepare doc
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                            leftMargin=0.35 * inch, rightMargin=0.35 * inch,
                            topMargin=0.35 * inch, bottomMargin=0.35 * inch)
    elements = []

    # Title block
    title_style = ParagraphStyle("title", parent=styles["Title"], alignment=1, fontSize=14, leading=16)
    title_text = (
        f"{safe_str(metadata.get('department_name', 'DEPARTMENT')).upper()}<br/>"
        f"{safe_str(metadata.get('date_range', '')).upper()}<br/>"
        f"{safe_str(metadata.get('report_title', 'ATTENDANCE REPORT')).upper()}"
    )
    elements.append(Paragraph(title_text, title_style))
    elements.append(Spacer(1, 0.12 * inch))

    # Metadata
    meta_style = ParagraphStyle("meta", parent=styles["Normal"], fontSize=9, leading=11)
    meta_block = (
        f"<b>Branch:</b> {safe_str(metadata.get('branch', 'N/A'))} &nbsp; "
        f"<b>Department:</b> {safe_str(metadata.get('department_specialization', 'N/A'))}<br/>"
        f"<b>Class:</b> {safe_str(metadata.get('class_name_division', 'N/A'))} &nbsp; "
        f"<b>Division:</b> {safe_str(metadata.get('division', 'N/A'))} &nbsp; "
        f"<b>Date:</b> {safe_str(metadata.get('date_range', 'N/A'))} &nbsp; "
        f"<b>Coordinator:</b> {safe_str(metadata.get('coordinator', ''))}"
    )
    elements.append(Paragraph(meta_block, meta_style))
    elements.append(Spacer(1, 0.12 * inch))

    # Build headers
    raw_headers = list(df.columns)
    cleaned_headers = [clean_subject_label(h) for h in raw_headers]
    hdr_style = ParagraphStyle("hdr", fontSize=8, leading=9, alignment=1)
    wrapped_headers = [Paragraph(h or "", hdr_style) for h in cleaned_headers]
    
    num_cols = len(cleaned_headers)

    # 1. Define the "invisible boundary" trigger
    COLUMN_THRESHOLD = 15
    apply_aggressive_wrapping = num_cols > COLUMN_THRESHOLD

    # 2. Dynamically set the width of the name column
    name_col_idx = 2
    try:
        header_texts = [h.text.strip() for h in wrapped_headers]
        name_col_idx = header_texts.index("Student Name")
    except ValueError:
        pass

    if apply_aggressive_wrapping:
        name_col_width = 0.8 * inch
    else:
        name_col_width = 1.65 * inch
    
    left_fixed = [0.45 * inch, 1.15 * inch, 1.65 * inch]  # Default widths
    left_fixed[name_col_idx] = name_col_width  # Overwrite with dynamic width

    # 3. Normalize data rows with selective wrapping
    left_align_style = ParagraphStyle("data_cell_left", parent=styles["Normal"], fontSize=8, leading=9, alignment=0)
    normalized_rows = []

    for row in df.values.tolist():
        row_list = []
        for i, cell_text in enumerate(list(row)):
            if i == name_col_idx:
                formatted_text = safe_str(cell_text)
                if apply_aggressive_wrapping:
                    formatted_text = formatted_text.replace(' ', '<br/>')
                row_list.append(Paragraph(formatted_text, left_align_style))
            elif i > name_col_idx: # Apply formatting to columns after Student Name (likely attendance)
                row_list.append(safe_str(format_attendance_value(cell_text)))
            else:
                row_list.append(safe_str(cell_text))
        
        if len(row_list) < num_cols:
            row_list += [""] * (num_cols - len(row_list))
        elif len(row_list) > num_cols:
            row_list = row_list[:num_cols]
        normalized_rows.append(row_list)

    table_data = [wrapped_headers] + normalized_rows

    # Column width engine
    page_w = landscape(letter)[0]
    available = page_w - doc.leftMargin - doc.rightMargin

    right_fixed_count = 4 if num_cols >= (len(left_fixed) + 4) else max(0, num_cols - len(left_fixed) - len(subject_details))
    right_fixed_width = 0.8 * inch

    num_subject_cols = max(1, num_cols - (len(left_fixed) + right_fixed_count))
    remaining_width = available - sum(left_fixed) - (right_fixed_count * right_fixed_width)
    subject_w = max(0.42 * inch, remaining_width / max(1, num_subject_cols))
    
    col_widths = list(left_fixed)
    col_widths += [subject_w] * num_subject_cols
    col_widths += [right_fixed_width] * right_fixed_count

    if len(col_widths) < num_cols:
        col_widths += [subject_w] * (num_cols - len(col_widths))
    col_widths = col_widths[:num_cols]

    # Build table
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(1, 1, 0, alpha=0.2)),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.black),
    ])

    # 4. Update highlighting logic
    min_att = metadata.get("min_attendance", 75)
    for subj, col_name in subject_details.items():
        possible_cols = [c for c in raw_headers if str(c).startswith(subj) or clean_subject_label(c) == subj]
        for pc in possible_cols:
            if pc not in raw_headers: continue
            col_idx = raw_headers.index(pc)
            
            for ridx, row in enumerate(normalized_rows, start=1):
                try:
                    cell_content = row[col_idx]
                    val_str = cell_content.text if isinstance(cell_content, Paragraph) else cell_content
                    val = float(safe_str(val_str))
                    if val < min_att:
                        tbl_style.add("BACKGROUND", (col_idx, ridx), (col_idx, ridx), colors.lightgrey)
                except (ValueError, IndexError):
                    pass

    table.setStyle(tbl_style)
    elements.append(table)
    elements.append(Spacer(1, 0.12 * inch))

    # Page break after main table
    elements.append(PageBreak())

    # Summary table
    subjects = list(subject_details.keys())
    valid_subjects = [s for s in subjects if s in df.columns or any(str(c).startswith(s) for c in df.columns)]
    summary_headers = ["Subject", "Students < 75%", "Students < 70%", "Students < 65%", "Students < 60%"]
    summary_rows = [summary_headers]
    for s in valid_subjects:
        col_key = None
        for c in df.columns:
            if str(c).startswith(s) or clean_subject_label(c) == s:
                col_key = c
                break
        if col_key is None:
            continue
        summary_rows.append([
            s,
            int((pd.to_numeric(df[col_key], errors="coerce") < 75).sum()),
            int((pd.to_numeric(df[col_key], errors="coerce") < 70).sum()),
            int((pd.to_numeric(df[col_key], errors="coerce") < 65).sum()),
            int((pd.to_numeric(df[col_key], errors="coerce") < 60).sum()),
        ])

    sum_hdr = ParagraphStyle("sumhdr", fontSize=8, alignment=1)
    sum_sub = ParagraphStyle("sumsub", fontSize=8, alignment=0)
    wrapped_summary = []
    for i, row in enumerate(summary_rows):
        new_row = []
        for j, cell in enumerate(row):
            if j == 0:
                if i == 0:
                    new_row.append(Paragraph(str(cell), sum_hdr))
                else:
                    new_row.append(Paragraph(str(cell), sum_sub))
            else:
                new_row.append(str(cell))
        wrapped_summary.append(new_row)

    summary_col_widths = [
        4.0 * inch, 1.5 * inch, 1.5 * inch, 1.5 * inch, 1.5 * inch
    ]

    summary_table = Table(wrapped_summary, colWidths=summary_col_widths, hAlign="LEFT")
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.black),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.12 * inch))

    # Page break after summary table
    elements.append(PageBreak())

    # Optional chart image
    if chart_image:
        try:
            rl_img = RLImage(chart_image)
            
            img_w, img_h = rl_img.imageWidth, rl_img.imageHeight
            if img_w <= 0 or img_h <= 0:
                raise ValueError("Invalid image dimensions")

            aspect = img_h / float(img_w)
            new_w = doc.width
            new_h = new_w * aspect

            if new_h > doc.height:
                new_h = doc.height
                new_w = new_h / aspect
            
            rl_img.drawWidth = new_w
            rl_img.drawHeight = new_h
            rl_img.hAlign = 'CENTER'
            elements.append(rl_img)
        except Exception:
            logger.exception("Failed to attach chart image to PDF")

    # Build and return buffer
    doc.build(elements)
    buffer.seek(0)
    return buffer


# -------------------------
# Excel generation (openpyxl)
# -------------------------
def get_adjusted_width(worksheet, column_cells) -> int:
    """Calculate a reasonable width for an openpyxl column based on text length."""
    max_length = 0
    for c in column_cells:
        try:
            if c.value is not None:
                length = len(str(c.value))
                if length > max_length:
                    max_length = length
        except Exception:
            pass
    # small fudge factor
    return max(8, min(50, max_length + 2))


def update_worksheet(worksheet, df: pd.DataFrame):
    """
    Auto-adjust column widths on an openpyxl worksheet.
    For subject columns we keep a fixed width to avoid very wide sheets.
    """
    for i, col in enumerate(worksheet.columns, start=1):
        col_letter = get_column_letter(i)
        # For subject columns we set a reasonable width, else auto adjust
        # We treat subject columns as those after 3rd and before last 3 columns
        try:
            if 3 < i <= df.shape[1] - 3:
                worksheet.column_dimensions[col_letter].width = 15
            else:
                worksheet.column_dimensions[col_letter].width = get_adjusted_width(worksheet, col)
        except Exception:
            # fallback
            worksheet.column_dimensions[col_letter].width = 15


def add_custom_header(ws, metadata):
    """Adds a dynamic, multi-line, centered, and bolded header to the worksheet."""
    # Safe fallbacks for metadata
    dept_name = metadata.get('department_name', 'DEPT OF COMPUTER SCIENCE & TECHNOLOGY')
    academic_year = metadata.get('academic_year', '2025-2026')
    semester = metadata.get('semester', 'Odd')
    report_title = metadata.get('report_title', 'ATTENDANCE MONITORING REPORT')
    branch = metadata.get('branch', 'MRU-School of Engineering')
    department_specialization = metadata.get('department_specialization', 'B.Tech (Hons.) in Computer Science Engineering with specializations in Gen AI')
    class_name_division = metadata.get('class_name_division', 'B.Tech CSE Gen AI Sem 1 | Division: All')
    date_range = metadata.get('date_range', '28/07/2025 to 19/09/2025 (2025-2026)')
    coordinator = metadata.get('coordinator', '')

    # Define header lines
    header_lines = [
        (dept_name, 'A1'),
        (f"Academic Year: {academic_year} - Semester: {semester}", 'A2'),
        (report_title, 'A3'),
        (f"Branch: {branch}", 'A4'),
        (f"Department: {department_specialization}", 'A5'),
        (f"Class Name: {class_name_division}", 'A6'),
        (f"Date: {date_range}", 'A7'),
        (f"Program Coordinator: {coordinator}", 'A8')
    ]

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')

    for line, cell_ref in header_lines:
        cell = ws[cell_ref]
        cell.value = line
        cell.font = bold_font
        cell.alignment = center_alignment
        # Merge cells from column A to the last column of the header
        ws.merge_cells(start_row=cell.row, start_column=1, end_row=cell.row, end_column=8)
def create_excel_file(df, subject_details, metadata, chart_image=None, report_color='#FFFF00'):
    """Create an excel file safely without corrupting fills."""
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage

    if "Roll No_duplicate" in df.columns and "Roll No" not in df.columns:
        df = df.rename(columns={"Roll No_duplicate": "Roll No"})

    wb = Workbook()
    ws = wb.active
    ws.title = metadata.get("monitoring_stage", "Report")
    
    # Add the custom header
    add_custom_header(ws, metadata)


    # Styles (valid only)
    vibrant_fill = PatternFill(start_color=report_color.replace("#", ""), end_color=report_color.replace("#", ""), fill_type="solid")
    grey_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
    light_blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

    # Header rows
    start_row = 10
    headers = list(df.columns)

    overall_idx = headers.index("Overall %age of all subjects from ERP report") + 1 \
                  if "Overall %age of all subjects from ERP report" in headers else 3

    # Write header row
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        if col_index <= overall_idx:
            cell.fill = vibrant_fill  # SAFE
        # DO NOT assign fill=None EVER

    # Write data
    min_attendance = metadata.get("min_attendance", 75)
    for r, row_data in enumerate(df.values.tolist(), start=start_row + 1):
        for c, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

            # Highlight subject columns
            if 4 <= c <= df.shape[1] - 3:
                formatted_val = format_attendance_value(val)
                cell.value = formatted_val
                try:
                    if float(val) < min_attendance:
                        cell.fill = grey_fill
                except:
                    pass

    # Summary table
    summary_start = start_row + len(df) + 3
    subjects = list(subject_details.keys())
    valid_subjects = [s for s in subjects if s in df.columns]

    summary_headers = ["Subject", "<75%", "<70%", "<65%", "<60%"]

    for i, header in enumerate(summary_headers, start=2):
        cell = ws.cell(row=summary_start, column=i, value=header)
        cell.font = Font(bold=True)
        cell.fill = light_blue_fill  # SAFE
        cell.border = thin_border

    for r, subject in enumerate(valid_subjects, start=summary_start + 1):
        ws.cell(row=r, column=2, value=subject).fill = light_blue_fill
        thresholds = [75, 70, 65, 60]
        for j, t in enumerate(thresholds, start=3):
            count = (df[subject] < t).sum()
            cell = ws.cell(row=r, column=j, value=int(count))
            cell.fill = light_blue_fill
            cell.border = thin_border

    # Autosize
    update_worksheet(ws, df)

    # Chart image
    if chart_image:
        img = XLImage(chart_image)
        img.anchor = f"B{summary_start + len(valid_subjects) + 3}"
        ws.add_image(img)

    # Save workbook
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out



# -------------------------
# HTML summary helper
# -------------------------
def generate_summary_table_html(df: pd.DataFrame, min_attendance: int = 75) -> str:
    """
    Return an HTML snippet representing summary counts for each subject.
    """
    subject_columns = [col for col in df.columns if col not in [
        "Sr No.", "Roll No", "Student Name", "Overall %age of all subjects from ERP report",
        "Roll No_duplicate", "Count of Courses with attendance below minimum attendance criteria",
        "Whether Critical"
    ]]

    summary_data = []
    for subject in subject_columns:
        summary_data.append({
            "Subject": subject,
            f"Below {min_attendance}%": int((pd.to_numeric(df[subject], errors="coerce") < min_attendance).sum()),
            "Below 70%": int((pd.to_numeric(df[subject], errors="coerce") < 70).sum()),
            "Below 65%": int((pd.to_numeric(df[subject], errors="coerce") < 65).sum()),
            "Below 60%": int((pd.to_numeric(df[subject], errors="coerce") < 60).sum()),
        })
    summary_df = pd.DataFrame(summary_data)
    return summary_df.to_html(classes="summary-table", index=False)


# -------------------------
# Chart generation
# -------------------------
def generate_chart_image(df: pd.DataFrame) -> BytesIO:
    """
    Generate a PNG Bar chart as BytesIO where x=courses and y=students below 75%.
    """
    subject_columns = [col for col in df.columns if col not in [
        "Sr No.", "Roll No", "Student Name", "Overall %age of all subjects from ERP report"
        ,"Count of Courses with attendance below minimum attendance criteria",
        "Whether Critical"
    ]]

    courses = subject_columns
    students_below_75 = [(pd.to_numeric(df[c], errors="coerce") < 75).sum() for c in courses]

    # Wrap long course names
    wrapped_courses = [textwrap.fill(course, 15) for course in courses]

    fig = Figure(figsize=(12, 6))
    ax = fig.subplots()
    bars = ax.bar(wrapped_courses, students_below_75)
    ax.set_title("Number of Students with Attendance Below 75% per Course")
    ax.set_xlabel("Courses")
    ax.set_ylabel("Number of Students below 75%")
    ax.tick_params(axis="x", rotation=45, labelsize=8)
    for b in bars:
        yval = b.get_height()
        ax.text(b.get_x() + b.get_width() / 2.0, yval, str(int(yval)), va="bottom", ha="center", fontsize=8)
    fig.tight_layout()

    buf = BytesIO()
    fig.savefig(buf, format="png")
    buf.seek(0)
    return buf

